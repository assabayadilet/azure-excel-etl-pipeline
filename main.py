"""
Azure Function: Excel → Azure SQL & Blob ETL Pipeline

Description:
    - Downloads an Excel file from Azure Blob Storage based on user email
    - Extracts data and images
    - Cleans and uploads data into an Azure SQL database
    - Archives and updates related images in Blob Storage
"""

import os
import logging
import urllib.parse
import json
from io import BytesIO
from datetime import datetime

import pytz
import pandas as pd
import azure.functions as func
from sqlalchemy import create_engine, text
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader


def main(req: func.HttpRequest) -> func.HttpResponse:
    """Main entry point for the Azure Function."""
    logging.info("Processing request for Excel upload and database update.")

    # --- User identification (e.g., email or user ID) ---
    user_email = req.params.get("email")
    logging.info(f"Request from user: {user_email}")

    # --- Initialize Azure connections ---
    blob_conn_str = os.getenv("AZURE_BLOB_STORAGE_KEY")
    container_name = os.getenv("AZURE_CONTAINER_NAME_KEY")
    db_conn_str = os.getenv("DB_CONNECTION_STRING_KEY")

    blob_service = BlobServiceClient.from_connection_string(blob_conn_str)
    container_client = blob_service.get_container_client(container_name)

    # --- Locate the target Excel file in Blob Storage ---
    blob_data = None
    for blob in container_client.list_blobs(name_starts_with="DATASET/Current"):
        if user_email and user_email in blob.name:
            logging.info(f"Found Excel file: {blob.name}")
            blob_data = container_client.download_blob(blob.name)
            break

    if not blob_data:
        msg = f"No file found for user {user_email}"
        logging.warning(msg)
        return func.HttpResponse(msg, status_code=404)

    # --- Load Excel file into DataFrame ---
    content = blob_data.content_as_bytes()
    df = pd.read_excel(content, sheet_name="Sheet1")

    # --- Data preprocessing ---
    tz = pytz.timezone("Asia/Tashkent")
    current_time = datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")

    # Remove unnecessary columns
    df = df.drop(columns=["Unused_Column"], errors="ignore")

    # Rename and add metadata
    df.rename(columns={"OldName": "DATASET_ID"}, inplace=True)
    df["Create_date"] = current_time
    df["Last_modified"] = current_time
    df["Created_by"] = user_email

    # Create image link reference
    df["Image_link"] = df["DATASET_ID"].apply(
        lambda x: f"/{container_name}/DATASET/{x}/{x}.png"
    )

    # Replace stray encoding artifacts
    df.replace({"Ð¢": "T", "Ðš": "K"}, regex=True, inplace=True)

    # --- Upload to Azure SQL ---
    params = urllib.parse.quote_plus(db_conn_str)
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")

    try:
        df.to_sql("DATASET", con=engine, if_exists="append", index=False)
        logging.info("Data successfully uploaded to Azure SQL.")
    except Exception as e:
        logging.error(f"Failed to upload data: {e}")
        return func.HttpResponse("Database upload failed.", status_code=500)

    # --- Process images embedded in Excel ---
    workbook = load_workbook(filename=BytesIO(content))
    worksheets = workbook.worksheets[4:]
    excluded_cells = ["A1", "K1", "L1", "M3", "I1"]

    for sheet in worksheets:
        image_loader = SheetImageLoader(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                if image_loader.image_in(cell.coordinate) and cell.coordinate not in excluded_cells:
                    try:
                        image = image_loader.get(cell.coordinate)
                        png_buffer = BytesIO()
                        image.save(png_buffer, format="PNG")
                        png_buffer.seek(0)

                        file_path = f"DATASET/{sheet.title}/{sheet.title}.png"
                        blob_client = container_client.get_blob_client(file_path)

                        # Archive existing image if found
                        if blob_client.exists():
                            archive_path = f"DATASET/{sheet.title}/Archive/{sheet.title}_{current_time}.png"
                            archive_client = blob_service.get_blob_client(container_name, archive_path)
                            archive_client.start_copy_from_url(blob_client.url)
                            blob_client.delete_blob()

                        blob_client.upload_blob(png_buffer, overwrite=True)
                        logging.info(f"Uploaded image: {file_path}")
                        break
                    except Exception as e:
                        logging.warning(f"Image processing failed: {e}")
                        continue

    return func.HttpResponse(json.dumps({"msg": "success"}), status_code=200)
